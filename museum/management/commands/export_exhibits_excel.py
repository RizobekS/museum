import os
from io import BytesIO

from django.core.management.base import BaseCommand
from django.conf import settings
from django.db import models

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

from PIL import Image as PILImage

from museum.models import Exhibit


class Command(BaseCommand):
    help = "Экспорт экспонатов в Excel с картинками (QR и single_image), с фильтром по блоку и оптимизацией изображений."

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            "-o",
            default="exhibits_export.xlsx",
            help="Путь к XLSX файлу (относительно BASE_DIR или абсолютный)",
        )
        parser.add_argument(
            "--only-with-photos",
            action="store_true",
            help="Выгружать только экспонаты, у которых есть QR или single_image",
        )
        parser.add_argument(
            "--block-slug",
            "-b",
            help="Фильтровать экспонаты по slug блока (MuseumBlock.slug), напр.: REN2",
        )
        parser.add_argument(
            "--no-images",
            action="store_true",
            default=False,
            help="Не встраивать картинки в Excel (только текст, файл будет значительно легче).",
        )
        parser.add_argument(
            "--thumb-size",
            type=int,
            default=400,
            help="Максимальный размер стороны превью в пикселях (по умолчанию 400). "
                 "Используется только если картинки встраиваются.",
        )

    def handle(self, *args, **options):
        output_path = options["output"]

        if not os.path.isabs(output_path):
            output_path = os.path.join(settings.BASE_DIR, output_path)

        qs = Exhibit.objects.select_related("block", "section")

        # Фильтр по блоку
        block_slug = options.get("block_slug")
        if block_slug:
            qs = qs.filter(block__slug=block_slug)
            self.stdout.write(self.style.WARNING(f"Фильтруем по блоку: {block_slug}"))

        # Фильтр по наличию фоток
        if options["only_with_photos"]:
            qs = qs.filter(
                models.Q(qr_code__isnull=False) | ~models.Q(single_image="")
            ).distinct()

        no_images = options.get("no_images", False)
        thumb_size = options.get("thumb_size", 400)

        if no_images:
            self.stdout.write(self.style.WARNING("Картинки НЕ будут встроены в Excel (режим --no-images)."))
        else:
            self.stdout.write(self.style.WARNING(
                f"Картинки будут встроены с уменьшением до {thumb_size} px по большей стороне."
            ))

        wb = Workbook()
        ws = wb.active
        ws.title = "Exhibits"

        # Заголовки колонок
        headers = [
            "block",
            "section",
            "title_ru",
            "title_uz",
            "title_en",
            "title_ar",
            "slug",
            "sub_title_ru",  # sub_title_ru + description_ru
            "sub_title_uz",  # sub_title_uz + description_uz
            "sub_title_en",  # sub_title_en + description_en
            "sub_title_ar",  # sub_title_ar + description_ar
            "qr_code",
            "single_image",
        ]

        # Записываем шапку
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Немного адекватной ширины колонок
        width_map = {
            "block": 20,
            "section": 25,
            "title_ru": 35,
            "title_uz": 35,
            "title_en": 35,
            "title_ar": 35,
            "slug": 25,
            "sub_title_ru": 80,
            "sub_title_uz": 80,
            "sub_title_en": 80,
            "sub_title_ar": 80,
            "qr_code": 20,
            "single_image": 25,
        }

        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width_map.get(header, 20)

        # Индексы колонок для картинок (для удобства)
        qr_col_idx = headers.index("qr_code") + 1
        single_img_col_idx = headers.index("single_image") + 1

        def combine(sub_title: str | None, description: str | None) -> str:
            """
            Объединяем sub_title и description в одну строку с переносом.
            Если одного из них нет, возвращаем то, что есть.
            """
            sub_title = sub_title or ""
            description = description or ""
            if sub_title and description:
                return f"{sub_title}\n\n{description}"
            return sub_title or description or ""

        def create_thumbnail_image(path: str, max_side: int):
            """
            Открываем оригинал, уменьшаем через Pillow, сохраняем в память (BytesIO)
            и создаём openpyxl Image из этого буфера.
            """
            try:
                with PILImage.open(path) as pil_img:
                    pil_img = pil_img.convert("RGBA")  # на всякий случай
                    pil_img.thumbnail((max_side, max_side), PILImage.LANCZOS)

                    buf = BytesIO()
                    # Excel нормально переваривает PNG
                    pil_img.save(buf, format="PNG")
                    buf.seek(0)
            except Exception as e:
                self.stderr.write(f"Не удалось подготовить превью для {path}: {e}")
                return None

            try:
                img = XLImage(buf)
            except Exception as e:
                self.stderr.write(f"Не удалось создать Excel-изображение для {path}: {e}")
                return None

            return img

        row_idx = 2
        count = qs.count()
        self.stdout.write(self.style.SUCCESS(f"Найдено экспонатов для экспорта: {count}"))

        for ex in qs:
            # Скомбинированные тексты по языкам
            text_ru = combine(ex.sub_title_ru, ex.description_ru)
            text_uz = combine(ex.sub_title_uz, ex.description_uz)
            text_en = combine(ex.sub_title_en, ex.description_en)
            text_ar = combine(ex.sub_title_ar, ex.description_ar)

            # Текстовые поля
            row_values = [
                ex.block.title_ru if ex.block else "",
                ex.section.title_ru if ex.section else "",
                ex.title_ru or "",
                ex.title_uz or "",
                ex.title_en or "",
                ex.title_ar or "",
                ex.slug or "",
                text_ru,
                text_uz,
                text_en,
                text_ar,
                "",  # qr_code (картинка)
                "",  # single_image (картинка)
            ]

            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True,
                )

            # Базовая высота строки
            ws.row_dimensions[row_idx].height = 40

            if not no_images:
                # Вставка QR-кода
                if ex.qr_code and ex.qr_code.name:
                    qr_path = os.path.join(settings.MEDIA_ROOT, ex.qr_code.name)
                    if os.path.exists(qr_path):
                        img = create_thumbnail_image(qr_path, thumb_size)
                        if img:
                            cell_addr = f"{get_column_letter(qr_col_idx)}{row_idx}"
                            ws.add_image(img, cell_addr)
                            ws.row_dimensions[row_idx].height = max(
                                ws.row_dimensions[row_idx].height, 120
                            )

                # Вставка single_image
                if ex.single_image and ex.single_image.name:
                    img_path = os.path.join(settings.MEDIA_ROOT, ex.single_image.name)
                    if os.path.exists(img_path):
                        img = create_thumbnail_image(img_path, thumb_size)
                        if img:
                            cell_addr = f"{get_column_letter(single_img_col_idx)}{row_idx}"
                            ws.add_image(img, cell_addr)
                            ws.row_dimensions[row_idx].height = max(
                                ws.row_dimensions[row_idx].height, 160
                            )

            row_idx += 1

        # Создаём директорию, если её нет
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        wb.save(output_path)
        self.stdout.write(self.style.SUCCESS(f"Экспорт завершён: {output_path}"))
